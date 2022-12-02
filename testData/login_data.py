import openpyxl as openpyxl


class getData:
    @staticmethod
    def get_data():

        excel_book = openpyxl.load_workbook("C:\\Users\\jakho\\Desktop\\SwagBucks.xlsx")

        sheet = excel_book.active

        loginData = {}

        # for i in range(1, sheet.max_row + 1):
        #     # if sheet.cell(row=i, column=i).value == user_name:
        #     for j in range(1, sheet.max_column + 1):
        #         loginData[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                loginData[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [loginData]
